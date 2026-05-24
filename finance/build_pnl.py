"""
Build the P&L template workbook for Ahmad's consulting business.
Run: python finance/build_pnl.py
Outputs: finance/pnl-template.xlsx
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Palette pulled from the website brand.
INK = "FF0F172A"
DEEP = "FF1E293B"
SLATE = "FF334155"
IVORY = "FFF7F1ED"
BLUSH = "FFE9D5D1"
ROSE = "FFD18C8C"
MAUVE = "FF8B5E5E"
LIGHT_RULE = "FFE5E5E0"

THIN = Side(border_style="thin", color=LIGHT_RULE)
BORDER_BOTTOM = Border(bottom=THIN)
BORDER_TOP = Border(top=Side(border_style="thin", color=SLATE))
BORDER_BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

EYEBROW = Font(name="Inter", size=9, color=MAUVE, bold=True)
HEADING = Font(name="Georgia", size=22, color=INK, bold=False, italic=False)
SUBHEAD = Font(name="Georgia", size=14, color=INK, italic=True)
LABEL = Font(name="Inter", size=10, color=INK)
LABEL_BOLD = Font(name="Inter", size=10, color=INK, bold=True)
SUM_FONT = Font(name="Inter", size=11, color=INK, bold=True)
CATEGORY_FONT = Font(name="Inter", size=10, color=MAUVE, bold=True)
INPUT_FONT = Font(name="Inter", size=10, color=DEEP)

FILL_HEAD = PatternFill("solid", fgColor=IVORY)
FILL_BLUSH = PatternFill("solid", fgColor=BLUSH)
FILL_SUM = PatternFill("solid", fgColor=IVORY)
FILL_INK = PatternFill("solid", fgColor=INK)

MONEY_FMT = '_-"SAR" * #,##0_-;[Red]_-"SAR" * -#,##0_-;_-"SAR" * "—"_-;_-@_-'
PCT_FMT = "0.0%"

MONTHS = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


# ----------------------------------------------------------------------------
# Categories — single source of truth used by both monthly + rollup sheets.
# ----------------------------------------------------------------------------

REVENUE_LINES = [
    ("Starter Brief — completed projects", "revenue"),
    ("Standard Build — completed projects", "revenue"),
    ("Studio Engagement — monthly retainers", "revenue"),
    ("Custom-scope projects", "revenue"),
    ("Advisory / hourly work", "revenue"),
    ("Other income", "revenue"),
]

DIRECT_COSTS = [
    ("Subcontractor fees", "direct"),
    ("Project-specific software / licences", "direct"),
    ("Stock assets, photography, fonts", "direct"),
    ("Other direct project costs", "direct"),
]

OPERATING_EXPENSES = [
    ("Software subscriptions (recurring tools)", "opex"),
    ("Cloud & hosting (Vercel, Supabase, Stripe fees)", "opex"),
    ("Domain & email infrastructure", "opex"),
    ("Equipment & hardware", "opex"),
    ("Internet & communications", "opex"),
    ("Office & workspace", "opex"),
    ("Travel & client meetings", "opex"),
    ("Marketing & advertising", "opex"),
    ("Professional services (legal, accounting)", "opex"),
    ("Training, books, conferences", "opex"),
    ("Banking & FX fees", "opex"),
    ("Other operating expenses", "opex"),
]

TAX_LINES = [
    ("VAT collected (output VAT, 15%)", "vat_out"),
    ("VAT paid on expenses (input VAT)", "vat_in"),
    ("Zakat provision", "zakat"),
]


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def set_widths(ws: Worksheet, widths: list[tuple[str, float]]) -> None:
    for col, w in widths:
        ws.column_dimensions[col].width = w


def write_cell(
    ws: Worksheet,
    coord: str,
    value,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    fmt: str | None = None,
    align: Alignment | None = None,
    border: Border | None = None,
) -> None:
    c = ws[coord]
    c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = align
    if border:
        c.border = border


# ----------------------------------------------------------------------------
# Sheet builders
# ----------------------------------------------------------------------------

def build_categories(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    set_widths(ws, [("A", 4), ("B", 50), ("C", 18), ("D", 50)])
    ws.row_dimensions[1].height = 40

    write_cell(ws, "B1", "Chart of accounts", font=HEADING)
    write_cell(ws, "B2", "The list every other sheet refers to. Add lines here, then add them on Monthly P&L.",
               font=Font(name="Inter", size=10, italic=True, color=SLATE))

    row = 5
    for section_name, lines in [
        ("Revenue", REVENUE_LINES),
        ("Direct costs (cost of sales)", DIRECT_COSTS),
        ("Operating expenses", OPERATING_EXPENSES),
        ("Tax & Zakat", TAX_LINES),
    ]:
        write_cell(ws, f"B{row}", section_name.upper(), font=CATEGORY_FONT)
        row += 1
        for label, kind in lines:
            write_cell(ws, f"B{row}", label, font=LABEL)
            write_cell(ws, f"C{row}", kind, font=Font(name="Inter", size=9, color=MAUVE, italic=True))
            row += 1
        row += 1


def build_monthly(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    set_widths(ws, [
        ("A", 4), ("B", 48), ("C", 18), ("D", 4), ("E", 48),
    ])
    ws.row_dimensions[1].height = 40

    write_cell(ws, "B1", "Monthly P&L", font=HEADING)
    write_cell(ws, "B2",
               "Fill the blue cells. Subtotals and totals calculate themselves.",
               font=Font(name="Inter", size=10, italic=True, color=SLATE))

    # Period inputs
    write_cell(ws, "B4", "MONTH", font=EYEBROW)
    write_cell(ws, "B5", "Month", font=LABEL_BOLD)
    write_cell(ws, "C5", "January 2026", font=INPUT_FONT, fill=PatternFill("solid", fgColor="FFFFFFFF"))
    write_cell(ws, "B6", "Currency", font=LABEL_BOLD)
    write_cell(ws, "C6", "SAR", font=INPUT_FONT)

    section_fills_for_input = PatternFill("solid", fgColor="FFFFFFFF")

    row = 9

    # REVENUE
    write_cell(ws, "B"+str(row), "REVENUE", font=CATEGORY_FONT)
    row += 1
    revenue_start = row
    for label, _ in REVENUE_LINES:
        write_cell(ws, f"B{row}", label, font=LABEL)
        write_cell(ws, f"C{row}", 0, font=INPUT_FONT, fmt=MONEY_FMT, fill=section_fills_for_input, border=BORDER_BOTTOM)
        row += 1
    revenue_end = row - 1

    # Gross revenue
    write_cell(ws, f"B{row}", "Gross revenue", font=SUM_FONT, fill=FILL_SUM)
    write_cell(ws, f"C{row}", f"=SUM(C{revenue_start}:C{revenue_end})",
               font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_SUM, border=BORDER_TOP)
    gross_rev_row = row
    row += 2

    # DIRECT COSTS
    write_cell(ws, f"B{row}", "DIRECT COSTS", font=CATEGORY_FONT)
    row += 1
    direct_start = row
    for label, _ in DIRECT_COSTS:
        write_cell(ws, f"B{row}", label, font=LABEL)
        write_cell(ws, f"C{row}", 0, font=INPUT_FONT, fmt=MONEY_FMT, fill=section_fills_for_input, border=BORDER_BOTTOM)
        row += 1
    direct_end = row - 1
    write_cell(ws, f"B{row}", "Total direct costs", font=SUM_FONT, fill=FILL_SUM)
    write_cell(ws, f"C{row}", f"=SUM(C{direct_start}:C{direct_end})",
               font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_SUM, border=BORDER_TOP)
    direct_total_row = row
    row += 2

    # Gross profit
    write_cell(ws, f"B{row}", "Gross profit", font=SUM_FONT, fill=FILL_BLUSH)
    write_cell(ws, f"C{row}", f"=C{gross_rev_row}-C{direct_total_row}",
               font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_BLUSH)
    gross_profit_row = row
    row += 1
    write_cell(ws, f"B{row}", "Gross margin %", font=LABEL)
    write_cell(ws, f"C{row}", f"=IFERROR(C{gross_profit_row}/C{gross_rev_row},0)",
               font=LABEL, fmt=PCT_FMT)
    row += 2

    # OPEX
    write_cell(ws, f"B{row}", "OPERATING EXPENSES", font=CATEGORY_FONT)
    row += 1
    opex_start = row
    for label, _ in OPERATING_EXPENSES:
        write_cell(ws, f"B{row}", label, font=LABEL)
        write_cell(ws, f"C{row}", 0, font=INPUT_FONT, fmt=MONEY_FMT, fill=section_fills_for_input, border=BORDER_BOTTOM)
        row += 1
    opex_end = row - 1
    write_cell(ws, f"B{row}", "Total operating expenses", font=SUM_FONT, fill=FILL_SUM)
    write_cell(ws, f"C{row}", f"=SUM(C{opex_start}:C{opex_end})",
               font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_SUM, border=BORDER_TOP)
    opex_total_row = row
    row += 2

    # EBITDA
    write_cell(ws, f"B{row}", "Operating profit (EBITDA)", font=SUM_FONT, fill=FILL_BLUSH)
    write_cell(ws, f"C{row}", f"=C{gross_profit_row}-C{opex_total_row}",
               font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_BLUSH)
    ebitda_row = row
    row += 1
    write_cell(ws, f"B{row}", "Operating margin %", font=LABEL)
    write_cell(ws, f"C{row}", f"=IFERROR(C{ebitda_row}/C{gross_rev_row},0)",
               font=LABEL, fmt=PCT_FMT)
    row += 2

    # TAX
    write_cell(ws, f"B{row}", "VAT & ZAKAT", font=CATEGORY_FONT)
    row += 1
    write_cell(ws, f"B{row}", "VAT collected (output VAT, 15% on taxable sales)", font=LABEL)
    write_cell(ws, f"C{row}", 0, font=INPUT_FONT, fmt=MONEY_FMT, fill=section_fills_for_input, border=BORDER_BOTTOM)
    vat_out_row = row
    row += 1
    write_cell(ws, f"B{row}", "VAT paid on expenses (input VAT)", font=LABEL)
    write_cell(ws, f"C{row}", 0, font=INPUT_FONT, fmt=MONEY_FMT, fill=section_fills_for_input, border=BORDER_BOTTOM)
    vat_in_row = row
    row += 1
    write_cell(ws, f"B{row}", "VAT payable to ZATCA", font=SUM_FONT)
    write_cell(ws, f"C{row}", f"=C{vat_out_row}-C{vat_in_row}", font=SUM_FONT, fmt=MONEY_FMT)
    row += 1
    write_cell(ws, f"B{row}", "Zakat provision (2.5% of zakat base — confirm with accountant)", font=LABEL)
    write_cell(ws, f"C{row}", 0, font=INPUT_FONT, fmt=MONEY_FMT, fill=section_fills_for_input, border=BORDER_BOTTOM)
    zakat_row = row
    row += 2

    # Net
    write_cell(ws, f"B{row}", "NET PROFIT (after zakat)", font=Font(name="Inter", size=12, color=IVORY, bold=True), fill=FILL_INK)
    write_cell(ws, f"C{row}", f"=C{ebitda_row}-C{zakat_row}",
               font=Font(name="Inter", size=12, color=IVORY, bold=True), fmt=MONEY_FMT, fill=FILL_INK)
    row += 1
    write_cell(ws, f"B{row}", "Net margin %", font=LABEL)
    write_cell(ws, f"C{row}", f"=IFERROR((C{ebitda_row}-C{zakat_row})/C{gross_rev_row},0)",
               font=LABEL, fmt=PCT_FMT)

    # Right-side notes column
    write_cell(ws, "E4", "NOTES & METHOD", font=EYEBROW)
    notes = [
        "1. Revenue is recognised when the invoice is issued, not when paid.",
        "2. VAT collected is NOT revenue — track it separately and remit to ZATCA quarterly.",
        "3. Stripe processing fees go under 'Cloud & hosting'.",
        "4. Subcontractor fees: include freelancers, designers, contract devs.",
        "5. Zakat applies to your net wealth base, not net profit — talk to an accountant.",
        "6. Keep all invoices and receipts. ZATCA can request them up to 5 years back.",
        "7. Numbers in red = outflows. Negative net profit means a loss month.",
        "",
        "To use this for another month: duplicate this sheet, rename it",
        "to the month, and edit the cells. The 12-Month Rollup sheet pulls",
        "from the monthly tabs by name — make sure to match exactly.",
    ]
    for i, note in enumerate(notes, start=5):
        write_cell(ws, f"E{i}", note, font=Font(name="Inter", size=10, color=SLATE))


def build_rollup(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    set_widths(ws, [("A", 4), ("B", 40)] + [(get_column_letter(i), 14) for i in range(3, 16)])
    ws.row_dimensions[1].height = 40

    write_cell(ws, "B1", "12-month rollup", font=HEADING)
    write_cell(ws, "B2",
               "A view of the full year. Fill the numbers here directly, or paste from each monthly sheet.",
               font=Font(name="Inter", size=10, italic=True, color=SLATE))

    # Month headers
    write_cell(ws, "B4", "LINE ITEM", font=EYEBROW)
    for i, m in enumerate(MONTHS, start=3):
        write_cell(ws, f"{get_column_letter(i)}4", m, font=EYEBROW, align=Alignment(horizontal="right"))
    write_cell(ws, "O4", "FULL YEAR", font=EYEBROW, align=Alignment(horizontal="right"))

    row = 5

    def write_input_line(label: str) -> int:
        write_cell(ws, f"B{row_local}", label, font=LABEL)
        for i in range(3, 15):
            col = get_column_letter(i)
            write_cell(ws, f"{col}{row_local}", 0,
                       font=INPUT_FONT, fmt=MONEY_FMT, border=BORDER_BOTTOM)
        write_cell(ws, f"O{row_local}", f"=SUM(C{row_local}:N{row_local})",
                   font=LABEL, fmt=MONEY_FMT, border=BORDER_BOTTOM)
        return row_local + 1

    # Revenue
    write_cell(ws, f"B{row}", "REVENUE", font=CATEGORY_FONT); row += 1
    rev_start = row
    for label, _ in REVENUE_LINES:
        row_local = row
        row = write_input_line(label)
    rev_end = row - 1
    write_cell(ws, f"B{row}", "Gross revenue", font=SUM_FONT, fill=FILL_SUM)
    for i in range(3, 16):
        col = get_column_letter(i)
        write_cell(ws, f"{col}{row}", f"=SUM({col}{rev_start}:{col}{rev_end})",
                   font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_SUM, border=BORDER_TOP)
    gross_rev = row
    row += 2

    # Direct costs
    write_cell(ws, f"B{row}", "DIRECT COSTS", font=CATEGORY_FONT); row += 1
    direct_start = row
    for label, _ in DIRECT_COSTS:
        row_local = row
        row = write_input_line(label)
    direct_end = row - 1
    write_cell(ws, f"B{row}", "Total direct costs", font=SUM_FONT, fill=FILL_SUM)
    for i in range(3, 16):
        col = get_column_letter(i)
        write_cell(ws, f"{col}{row}", f"=SUM({col}{direct_start}:{col}{direct_end})",
                   font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_SUM, border=BORDER_TOP)
    direct_total = row
    row += 2

    # Gross profit
    write_cell(ws, f"B{row}", "Gross profit", font=SUM_FONT, fill=FILL_BLUSH)
    for i in range(3, 16):
        col = get_column_letter(i)
        write_cell(ws, f"{col}{row}", f"={col}{gross_rev}-{col}{direct_total}",
                   font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_BLUSH)
    gross_profit = row
    row += 2

    # Opex
    write_cell(ws, f"B{row}", "OPERATING EXPENSES", font=CATEGORY_FONT); row += 1
    opex_start = row
    for label, _ in OPERATING_EXPENSES:
        row_local = row
        row = write_input_line(label)
    opex_end = row - 1
    write_cell(ws, f"B{row}", "Total operating expenses", font=SUM_FONT, fill=FILL_SUM)
    for i in range(3, 16):
        col = get_column_letter(i)
        write_cell(ws, f"{col}{row}", f"=SUM({col}{opex_start}:{col}{opex_end})",
                   font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_SUM, border=BORDER_TOP)
    opex_total = row
    row += 2

    # EBITDA
    write_cell(ws, f"B{row}", "Operating profit (EBITDA)", font=SUM_FONT, fill=FILL_BLUSH)
    for i in range(3, 16):
        col = get_column_letter(i)
        write_cell(ws, f"{col}{row}", f"={col}{gross_profit}-{col}{opex_total}",
                   font=SUM_FONT, fmt=MONEY_FMT, fill=FILL_BLUSH)
    ebitda = row
    row += 2

    # Zakat
    write_cell(ws, f"B{row}", "Zakat provision", font=LABEL)
    for i in range(3, 15):
        col = get_column_letter(i)
        write_cell(ws, f"{col}{row}", 0,
                   font=INPUT_FONT, fmt=MONEY_FMT, border=BORDER_BOTTOM)
    write_cell(ws, f"O{row}", f"=SUM(C{row}:N{row})",
               font=LABEL, fmt=MONEY_FMT, border=BORDER_BOTTOM)
    zakat = row
    row += 2

    # Net
    write_cell(ws, f"B{row}", "NET PROFIT",
               font=Font(name="Inter", size=12, color=IVORY, bold=True),
               fill=FILL_INK)
    for i in range(3, 16):
        col = get_column_letter(i)
        write_cell(ws, f"{col}{row}", f"={col}{ebitda}-{col}{zakat}",
                   font=Font(name="Inter", size=12, color=IVORY, bold=True),
                   fmt=MONEY_FMT, fill=FILL_INK)
    net = row
    row += 1
    write_cell(ws, f"B{row}", "Net margin %", font=LABEL)
    for i in range(3, 16):
        col = get_column_letter(i)
        write_cell(ws, f"{col}{row}", f"=IFERROR({col}{net}/{col}{gross_rev},0)",
                   font=LABEL, fmt=PCT_FMT)
    row += 2

    # KPIs
    write_cell(ws, f"B{row}", "KPIs (full year)", font=EYEBROW); row += 1
    write_cell(ws, f"B{row}", "Average monthly revenue", font=LABEL)
    write_cell(ws, f"O{row}", f"=IFERROR(O{gross_rev}/COUNTIF(C{gross_rev}:N{gross_rev},\">0\"),0)",
               font=LABEL, fmt=MONEY_FMT); row += 1
    write_cell(ws, f"B{row}", "Best month (revenue)", font=LABEL)
    write_cell(ws, f"O{row}", f"=MAX(C{gross_rev}:N{gross_rev})",
               font=LABEL, fmt=MONEY_FMT); row += 1
    write_cell(ws, f"B{row}", "Average gross margin %", font=LABEL)
    write_cell(ws, f"O{row}", f"=IFERROR(O{gross_profit}/O{gross_rev},0)",
               font=LABEL, fmt=PCT_FMT); row += 1
    write_cell(ws, f"B{row}", "Average net margin %", font=LABEL)
    write_cell(ws, f"O{row}", f"=IFERROR(O{net}/O{gross_rev},0)",
               font=LABEL, fmt=PCT_FMT); row += 1


def build_readme(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    set_widths(ws, [("A", 4), ("B", 90)])
    ws.row_dimensions[1].height = 40

    write_cell(ws, "B1", "P&L template — how to use", font=HEADING)

    lines = [
        ("", None),
        ("Three working sheets:", LABEL_BOLD),
        ("  • Categories — the chart of accounts every other sheet pulls from.", LABEL),
        ("  • Monthly P&L — one month at a time. Duplicate for each month.", LABEL),
        ("  • 12-Month Rollup — the year on one page, by month and full-year total.", LABEL),
        ("", None),
        ("How to use it each month", LABEL_BOLD),
        ("  1. Duplicate the 'Monthly P&L' tab and rename it to the month (e.g. '2026-05').", LABEL),
        ("  2. Update the Month label cell (C5).", LABEL),
        ("  3. Fill the white input cells with the totals from your bank statements,", LABEL),
        ("     invoicing system, and expense receipts.", LABEL),
        ("  4. Copy the row totals into the matching month column on '12-Month Rollup'.", LABEL),
        ("", None),
        ("Conventions used here", LABEL_BOLD),
        ("  • All amounts in SAR. If you bill a client in USD, convert at the rate", LABEL),
        ("    on the invoice date and keep the original USD figure in the description.", LABEL),
        ("  • Revenue is recognised when the invoice is issued, not when paid.", LABEL),
        ("  • VAT collected from clients is a liability, not revenue. Track it under", LABEL),
        ("    the VAT & Zakat section and remit to ZATCA quarterly.", LABEL),
        ("  • Zakat is calculated on net wealth, not profit. The line here is a", LABEL),
        ("    monthly provision; talk to your accountant to true it up at year-end.", LABEL),
        ("", None),
        ("What this template is not", LABEL_BOLD),
        ("  • Not an accounting system. Use it alongside ZATCA-compliant invoicing", LABEL),
        ("    (e.g. Mahasaba, Wafeq, or Zoho Books with the KSA add-on).", LABEL),
        ("  • Not a substitute for an accountant. At year-end, hand this to your", LABEL),
        ("    accountant alongside bank statements; they will prepare the formal", LABEL),
        ("    financials and zakat/tax returns.", LABEL),
        ("  • Not a cash-flow statement. This is accrual-style: invoices issued, not", LABEL),
        ("    cash received. Add a separate cash sheet if you need to track timing.", LABEL),
        ("", None),
    ]
    for i, (text, font) in enumerate(lines, start=3):
        if text:
            write_cell(ws, f"B{i}", text, font=font or Font(name="Inter", size=10, color=SLATE))


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main() -> None:
    wb = Workbook()
    # Default sheet
    ws_readme = wb.active
    ws_readme.title = "Read me"
    build_readme(ws_readme)

    ws_monthly = wb.create_sheet("Monthly P&L")
    build_monthly(ws_monthly)

    ws_rollup = wb.create_sheet("12-Month Rollup")
    build_rollup(ws_rollup)

    ws_categories = wb.create_sheet("Categories")
    build_categories(ws_categories)

    out = Path(__file__).parent / "pnl-template.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
